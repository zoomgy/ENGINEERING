import openpyxl as xl

wb = xl.load_workbook("transactions.xlsx")

sheet = wb['Sheet1']

for i in range(2,sheet.max_row + 1):
    original_cell = sheet.cell(i,3)
    corrected_cell = sheet.cell(i,4)
    corrected_cell.value = original_cell.value * 0.9
    print(corrected_cell.value)

wb.save("trasactions2.xlsx")